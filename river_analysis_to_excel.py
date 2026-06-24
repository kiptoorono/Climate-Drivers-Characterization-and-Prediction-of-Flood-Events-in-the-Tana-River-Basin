import pandas as pd
import numpy as np
from scipy import stats
from collections import Counter
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings('ignore')

# ===== CONFIGURATION =====

EXCEL_FILE = 'TanaRiver Flow.xlsx'
SHEET_NAME = 'Rawdata'
WINDOW_SIZE = 14

THRESHOLDS = {
    'Bura': 721,
    'Galole': 723,
    'Garsen': 1723
}

STATIONS = ['Bura', 'Galole', 'Garsen']
OUTPUT_FILE = 'Flood_Analysis_Results_Seasonality.xlsx'

# Kenya's Bimodal Rainfall Seasonality
LONG_RAINS = [4, 5, 6]        # April, May, June
SHORT_RAINS = [10, 11, 12]    # October, November, December

# ===== FORMATTING UTILITIES =====

def format_header_row(ws, row_num, num_cols):
    """Format header row with styling"""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def format_data_cells(ws, start_row, end_row, num_cols):
    """Format data cells with borders and alignment"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')

# ===== SEASONALITY CLASSIFICATION =====

def classify_season(month):
    """Classify month into season: Long Rains, Short Rains, or Dry"""
    if month in LONG_RAINS:
        return 'Long Rains (Apr-Jun)'
    elif month in SHORT_RAINS:
        return 'Short Rains (Oct-Dec)'
    else:
        return 'Dry Season'

# ===== YEAR-AWARE ROLLING MAX =====

def year_aware_rolling_max(series, dates, year_col, window=14):
    """Calculate 14-day rolling max that respects year boundaries."""
    result = []
    series_vals = series.values
    year_vals = year_col.values

    for i in range(len(series_vals)):
        current_year = year_vals[i]
        year_indices = np.where(year_vals == current_year)[0]
        year_start_idx = year_indices[0] if len(year_indices) > 0 else 0
        lookback_idx = max(0, i - window + 1)
        start_idx = max(year_start_idx, lookback_idx)
        window_data = series_vals[start_idx:i+1]
        
        if len(window_data) > 0:
            max_val = np.max(window_data)
            if not np.isnan(max_val):
                result.append(int(max_val))
            else:
                result.append(np.nan)
        else:
            result.append(np.nan)

    return result

# ===== IDENTIFY FLOOD EVENTS =====

def identify_flood_events(rolling_max_series, threshold):
    """Identify flood events: start when rolling_max > threshold, end when <= threshold."""
    event_numbers = []
    current_event = 0
    in_event = False

    for value in rolling_max_series:
        if pd.isna(value):
            event_numbers.append(0)
            continue

        if value > threshold:
            if not in_event:
                current_event += 1
                in_event = True
            event_numbers.append(current_event)
        else:
            in_event = False
            event_numbers.append(0)

    return event_numbers

# ===== MARK EVENT DATES =====

def mark_event_dates(event_numbers, dates):
    """Mark event start/end dates and duration."""
    start_dates = [None] * len(event_numbers)
    end_dates = [None] * len(event_numbers)
    durations = [None] * len(event_numbers)

    current_event = None
    event_start_idx = None
    event_start_date = None

    for idx, event_num in enumerate(event_numbers):
        if event_num > 0:
            if event_num != current_event:
                current_event = event_num
                event_start_idx = idx
                event_start_date = dates.iloc[idx]
                start_dates[idx] = event_start_date
        else:
            if current_event is not None:
                end_date = dates.iloc[idx - 1]
                duration = (end_date - event_start_date).days + 1
                end_dates[idx - 1] = end_date
                durations[idx - 1] = duration
                current_event = None

    if current_event is not None:
        end_date = dates.iloc[len(dates) - 1]
        duration = (end_date - event_start_date).days + 1
        end_dates[len(dates) - 1] = end_date
        durations[len(dates) - 1] = duration

    return start_dates, end_dates, durations

# ===== CALCULATE EVENT BLOCK MAX =====

def calculate_event_block_max(event_numbers, rolling_values):
    """Record highest rolling window value at event end date."""
    event_block_max = [None] * len(event_numbers)
    event_to_max = {}

    for idx, event_num in enumerate(event_numbers):
        if event_num > 0 and not pd.isna(rolling_values[idx]):
            if event_num not in event_to_max:
                event_to_max[event_num] = rolling_values[idx]
            else:
                event_to_max[event_num] = max(event_to_max[event_num], rolling_values[idx])

    for idx, event_num in enumerate(event_numbers):
        if event_num > 0:
            is_event_end = (idx == len(event_numbers) - 1) or (event_numbers[idx + 1] != event_num)
            if is_event_end:
                event_block_max[idx] = event_to_max.get(event_num, None)

    return event_block_max

# ===== LOAD & PREPARE DATA =====

print("=" * 90)
print("FLOOD EVENT ANALYSIS WITH BIMODAL SEASONALITY - EXCEL EXPORT")
print("=" * 90)
print("\\n[STEP 1] Loading data...")

raw_data = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, header=1, usecols=[0, 1, 2, 3])
raw_data.columns = ['date', 'Bura', 'Galole', 'Garsen']
raw_data['date'] = pd.to_datetime(raw_data['date'], errors='coerce')

for col in STATIONS:
    raw_data[col] = pd.to_numeric(raw_data[col], errors='coerce')

raw_data['year'] = raw_data['date'].dt.year
raw_data['month'] = raw_data['date'].dt.month
raw_data['season'] = raw_data['month'].apply(classify_season)
raw_data = raw_data.dropna(subset=['date']).reset_index(drop=True)

print(f"✓ Data loaded: {raw_data['date'].min()} to {raw_data['date'].max()}")
print(f"✓ Total records: {len(raw_data)}")
print(f"✓ Years covered: {raw_data['year'].min()} to {raw_data['year'].max()}")

# ===== PROCESS EACH STATION =====

results = {}

for station in STATIONS:
    print(f"\\n[STEP 2] Processing {station} station...")
    
    threshold = THRESHOLDS[station]
    rolling_max = year_aware_rolling_max(raw_data[station], raw_data['date'], raw_data['year'], WINDOW_SIZE)
    events = identify_flood_events(rolling_max, threshold)
    start_dates, end_dates, durations = mark_event_dates(events, raw_data['date'])
    block_max = calculate_event_block_max(events, rolling_max)
    
    results[station] = {
        'rolling_max': rolling_max,
        'events': events,
        'start_dates': start_dates,
        'end_dates': end_dates,
        'durations': durations,
        'block_max': block_max,
        'threshold': threshold
    }
    
    print(f"✓ {station}: {len([e for e in events if e > 0])} events identified")

# ===== CREATE EVENT DATAFRAMES WITH SEASONALITY =====

for station in STATIONS:
    events = results[station]['events']
    durations = results[station]['durations']
    block_max = results[station]['block_max']
    start_dates = results[station]['start_dates']
    
    # Map event number to its start date since start_dates is only populated at the event's start index
    event_start_dates = {}
    for idx, event_num in enumerate(events):
        if event_num > 0 and start_dates[idx] is not None:
            event_start_dates[event_num] = start_dates[idx]
            
    event_data_list = []
    for idx, event_num in enumerate(events):
        if event_num > 0 and durations[idx] is not None:
            start_date = event_start_dates.get(event_num, None)
            event_data_list.append({
                'event_num': event_num,
                'start_date': start_date,
                'year': start_date.year if start_date else raw_data['year'].iloc[idx],
                'month': start_date.month if start_date else None,
                'season': classify_season(start_date.month) if start_date else None,
                'duration': durations[idx],
                'peak': block_max[idx]
            })
    
    event_df = pd.DataFrame(event_data_list).drop_duplicates(subset=['event_num'])
    results[station]['event_df'] = event_df

# ===== CREATE EXCEL WORKBOOK =====

print(f"\\n[STEP 3] Creating Excel file: {OUTPUT_FILE}...")
wb = Workbook()
wb.remove(wb.active)

# ===== SHEET 1: EVENT FREQUENCY ANALYSIS =====

ws1 = wb.create_sheet("1. Event Frequency")
row = 1

ws1.cell(row=row, column=1).value = "ANALYSIS 1: EVENT FREQUENCY & CHARACTERIZATION"
ws1.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws1.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws1.merge_cells(f'A{row}:F{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    threshold = results[station]['threshold']
    
    ws1.cell(row=row, column=1).value = f"{station.upper()} STATION"
    ws1.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws1.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws1.merge_cells(f'A{row}:F{row}')
    row += 1
    
    years_coverage = raw_data['year'].max() - raw_data['year'].min() + 1
    metrics = {
        'Total Flood Events': len(event_df),
        'Years Covered': years_coverage,
        'Average Events/Year': f"{len(event_df) / years_coverage:.2f}",
        'Mean Duration (days)': f"{event_df['duration'].mean():.2f}",
        'Median Duration (days)': f"{event_df['duration'].median():.2f}",
        'Min Duration (days)': event_df['duration'].min(),
        'Max Duration (days)': event_df['duration'].max(),
        'Mean Peak Level': f"{event_df['peak'].mean():.2f}",
        'Median Peak Level': f"{event_df['peak'].median():.2f}",
        'Min Peak Level': f"{event_df['peak'].min():.2f}",
        'Max Peak Level': f"{event_df['peak'].max():.2f}",
        'Std Dev Peak Level': f"{event_df['peak'].std():.2f}",
        'Threshold': threshold,
        'Mean Exceedance': f"{(event_df['peak'] - threshold).mean():.2f}",
        'Max Exceedance': f"{(event_df['peak'] - threshold).max():.2f}"
    }
    
    for key, value in metrics.items():
        ws1.cell(row=row, column=1).value = key
        ws1.cell(row=row, column=2).value = value
        row += 1
    
    row += 1

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 20

# ===== SHEET 2: DECADE TREND ANALYSIS =====

ws2 = wb.create_sheet("2. Decade Trends")
row = 1

ws2.cell(row=row, column=1).value = "ANALYSIS 2: TREND ANALYSIS (Events/Decade)"
ws2.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws2.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws2.merge_cells(f'A{row}:D{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    event_df['decade'] = (event_df['year'] // 10) * 10
    decade_counts = event_df.groupby('decade').size()
    decade_durations = event_df.groupby('decade')['duration'].mean()
    decade_peaks = event_df.groupby('decade')['peak'].mean()
    
    decade_summary = pd.DataFrame({
        'Decade': [f"{int(d)}s" for d in decade_counts.index],
        'Event Count': decade_counts.values,
        'Avg Duration (days)': decade_durations.values,
        'Avg Peak Level': decade_peaks.values
    })
    
    ws2.cell(row=row, column=1).value = f"{station.upper()}"
    ws2.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws2.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws2.merge_cells(f'A{row}:D{row}')
    row += 1
    
    headers = ['Decade', 'Event Count', 'Avg Duration (days)', 'Avg Peak Level']
    for col, header in enumerate(headers, 1):
        ws2.cell(row=row, column=col).value = header
    format_header_row(ws2, row, len(headers))
    row += 1
    
    for idx, r in decade_summary.iterrows():
        ws2.cell(row=row, column=1).value = r['Decade']
        ws2.cell(row=row, column=2).value = r['Event Count']
        ws2.cell(row=row, column=3).value = round(r['Avg Duration (days)'], 2)
        ws2.cell(row=row, column=4).value = round(r['Avg Peak Level'], 2)
        row += 1
    
    row += 2

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 18

# ===== SHEET 3: SEASONALITY (ENHANCED) =====

ws3 = wb.create_sheet("3. Seasonality")
row = 1

ws3.cell(row=row, column=1).value = "ANALYSIS 3: SEASONALITY (Month of Event Start)"
ws3.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws3.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws3.merge_cells(f'A{row}:C{row}')
row += 2

month_names = {1:'January', 2:'February', 3:'March', 4:'April', 5:'May', 6:'June',
    7:'July', 8:'August', 9:'September', 10:'October', 11:'November', 12:'December'}

for station in STATIONS:
    events = results[station]['events']
    start_dates = results[station]['start_dates']
    
    months_with_events = []
    for idx, date in enumerate(start_dates):
        if date is not None:
            months_with_events.append(date.month)
    
    month_counts = Counter(months_with_events)
    
    seasonality_data = []
    for month in range(1, 13):
        count = month_counts.get(month, 0)
        seasonality_data.append({'Month': month_names[month], 'Count': count})
    
    seasonality_df = pd.DataFrame(seasonality_data)
    
    ws3.cell(row=row, column=1).value = f"{station.upper()}"
    ws3.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws3.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws3.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws3.cell(row=row, column=1).value = 'Month'
    ws3.cell(row=row, column=2).value = 'Event Count'
    format_header_row(ws3, row, 2)
    row += 1
    
    for idx, r in seasonality_df.iterrows():
        ws3.cell(row=row, column=1).value = r['Month']
        ws3.cell(row=row, column=2).value = r['Count']
        row += 1
    
    row += 2

ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 15

# ===== SHEET 4: SPATIAL COMPARISON =====

ws4 = wb.create_sheet("4. Spatial Comparison")
row = 1

ws4.cell(row=row, column=1).value = "ANALYSIS 4: SPATIAL COMPARISON (3 Stations)"
ws4.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws4.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws4.merge_cells(f'A{row}:G{row}')
row += 2

summary_table = []
for station in STATIONS:
    event_df = results[station]['event_df']
    years_coverage = raw_data['year'].max() - raw_data['year'].min() + 1
    summary_table.append({
        'Station': station,
        'Total Events': len(event_df),
        'Events/Year': len(event_df) / years_coverage,
        'Avg Duration (days)': event_df['duration'].mean(),
        'Avg Peak': event_df['peak'].mean(),
        'Max Peak': event_df['peak'].max(),
        'Threshold': results[station]['threshold']
    })

summary_df = pd.DataFrame(summary_table)

headers = list(summary_df.columns)
for col, header in enumerate(headers, 1):
    ws4.cell(row=row, column=col).value = header
format_header_row(ws4, row, len(headers))
row += 1

for idx, r in summary_df.iterrows():
    for col, header in enumerate(headers, 1):
        value = r[header]
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            value = round(value, 2)
        ws4.cell(row=row, column=col).value = value
    row += 1

row += 2
ws4.cell(row=row, column=1).value = "Peak Values Correlation Between Stations"
ws4.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
ws4.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws4.merge_cells(f'A{row}:C{row}')
row += 1

correlation_data = []
for i, st1 in enumerate(STATIONS):
    for st2 in STATIONS[i+1:]:
        valid_indices = [(idx, results[st1]['block_max'][idx], results[st2]['block_max'][idx])
            for idx in range(len(results[st1]['block_max']))
            if results[st1]['block_max'][idx] is not None and results[st2]['block_max'][idx] is not None]
        
        if len(valid_indices) > 1:
            p1 = [x[1] for x in valid_indices]
            p2 = [x[2] for x in valid_indices]
            corr = np.corrcoef(p1, p2)[0, 1]
            correlation_data.append({'Comparison': f"{st1} vs {st2}", 'Correlation (r)': corr})

corr_df = pd.DataFrame(correlation_data)
for col, header in enumerate(['Comparison', 'Correlation (r)'], 1):
    ws4.cell(row=row, column=col).value = header
format_header_row(ws4, row, 2)
row += 1

for idx, r in corr_df.iterrows():
    ws4.cell(row=row, column=1).value = r['Comparison']
    ws4.cell(row=row, column=2).value = round(r['Correlation (r)'], 3)
    row += 1

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws4.column_dimensions[col].width = 18

# ===== SHEET 5: ANNUAL & POT =====

ws5 = wb.create_sheet("5. Annual & POT")
row = 1

ws5.cell(row=row, column=1).value = "ANALYSIS 5: ANNUAL MAXIMUM & PEAKS OVER THRESHOLD (POT)"
ws5.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws5.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws5.merge_cells(f'A{row}:D{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    annual_max = event_df.groupby('year')['peak'].max()
    all_peaks = event_df['peak'].values
    
    ws5.cell(row=row, column=1).value = f"{station.upper()}"
    ws5.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws5.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws5.merge_cells(f'A{row}:D{row}')
    row += 1
    
    metrics = {
        'Annual Maximum - Mean': f"{annual_max.mean():.2f}",
        'Annual Maximum - Median': f"{annual_max.median():.2f}",
        'Annual Maximum - Std Dev': f"{annual_max.std():.2f}",
        'POT Count': len(all_peaks),
        'POT Mean': f"{all_peaks.mean():.2f}",
        'POT Median': f"{np.median(all_peaks):.2f}",
        'POT Std Dev': f"{all_peaks.std():.2f}"
    }
    
    for key, value in metrics.items():
        ws5.cell(row=row, column=1).value = key
        ws5.cell(row=row, column=2).value = value
        row += 1
    
    row += 2

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 20

# ========== NEW SHEETS: SEASONAL ANALYSIS ==========

# ===== SHEET 6: LONG RAINS ANALYSIS =====

ws6 = wb.create_sheet("6. Long Rains")
row = 1

ws6.cell(row=row, column=1).value = "ANALYSIS 6: LONG RAINS FLOODING (April-June)"
ws6.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws6.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws6.merge_cells(f'A{row}:G{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    long_rains_events = event_df[event_df['season'] == 'Long Rains (Apr-Jun)']
    
    ws6.cell(row=row, column=1).value = f"{station.upper()}"
    ws6.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws6.cell(row=row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws6.merge_cells(f'A{row}:G{row}')
    row += 1
    
    if len(long_rains_events) > 0:
        metrics = {
            'Events in Long Rains': len(long_rains_events),
            'Percentage of Annual Events': f"{(len(long_rains_events) / len(event_df) * 100):.1f}%",
            'Mean Duration (days)': f"{long_rains_events['duration'].mean():.2f}",
            'Median Duration (days)': f"{long_rains_events['duration'].median():.2f}",
            'Mean Peak Level': f"{long_rains_events['peak'].mean():.2f}",
            'Median Peak Level': f"{long_rains_events['peak'].median():.2f}",
            'Max Peak Level': f"{long_rains_events['peak'].max():.2f}",
            'Min Peak Level': f"{long_rains_events['peak'].min():.2f}",
            'Std Dev Peak': f"{long_rains_events['peak'].std():.2f}"
        }
    else:
        metrics = {
            'Events in Long Rains': 0,
            'Percentage of Annual Events': '0%',
            'No flooding events detected in this season': ''
        }
    
    for key, value in metrics.items():
        ws6.cell(row=row, column=1).value = key
        ws6.cell(row=row, column=2).value = value
        row += 1
    
    row += 2

ws6.column_dimensions['A'].width = 35
ws6.column_dimensions['B'].width = 20

# ===== SHEET 7: SHORT RAINS ANALYSIS =====

ws7 = wb.create_sheet("7. Short Rains")
row = 1

ws7.cell(row=row, column=1).value = "ANALYSIS 7: SHORT RAINS FLOODING (October-December)"
ws7.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws7.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws7.merge_cells(f'A{row}:G{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    short_rains_events = event_df[event_df['season'] == 'Short Rains (Oct-Dec)']
    
    ws7.cell(row=row, column=1).value = f"{station.upper()}"
    ws7.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws7.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws7.merge_cells(f'A{row}:G{row}')
    row += 1
    
    if len(short_rains_events) > 0:
        metrics = {
            'Events in Short Rains': len(short_rains_events),
            'Percentage of Annual Events': f"{(len(short_rains_events) / len(event_df) * 100):.1f}%",
            'Mean Duration (days)': f"{short_rains_events['duration'].mean():.2f}",
            'Median Duration (days)': f"{short_rains_events['duration'].median():.2f}",
            'Mean Peak Level': f"{short_rains_events['peak'].mean():.2f}",
            'Median Peak Level': f"{short_rains_events['peak'].median():.2f}",
            'Max Peak Level': f"{short_rains_events['peak'].max():.2f}",
            'Min Peak Level': f"{short_rains_events['peak'].min():.2f}",
            'Std Dev Peak': f"{short_rains_events['peak'].std():.2f}"
        }
    else:
        metrics = {
            'Events in Short Rains': 0,
            'Percentage of Annual Events': '0%',
            'No flooding events detected in this season': ''
        }
    
    for key, value in metrics.items():
        ws7.cell(row=row, column=1).value = key
        ws7.cell(row=row, column=2).value = value
        row += 1
    
    row += 2

ws7.column_dimensions['A'].width = 35
ws7.column_dimensions['B'].width = 20

# ===== SHEET 8: SEASONAL COMPARISON =====

ws8 = wb.create_sheet("8. Seasonal Comparison")
row = 1

ws8.cell(row=row, column=1).value = "ANALYSIS 8: SEASONAL COMPARISON - Which Season Drives Flooding?"
ws8.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws8.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws8.merge_cells(f'A{row}:I{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    long_rains = event_df[event_df['season'] == 'Long Rains (Apr-Jun)']
    short_rains = event_df[event_df['season'] == 'Short Rains (Oct-Dec)']
    dry_season = event_df[event_df['season'] == 'Dry Season']
    
    ws8.cell(row=row, column=1).value = f"{station.upper()}"
    ws8.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws8.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws8.merge_cells(f'A{row}:I{row}')
    row += 1
    
    ws8.cell(row=row, column=1).value = 'Season'
    ws8.cell(row=row, column=2).value = 'Event Count'
    ws8.cell(row=row, column=3).value = '% of Total'
    ws8.cell(row=row, column=4).value = 'Mean Duration'
    ws8.cell(row=row, column=5).value = 'Median Duration'
    ws8.cell(row=row, column=6).value = 'Mean Peak'
    ws8.cell(row=row, column=7).value = 'Median Peak'
    ws8.cell(row=row, column=8).value = 'Max Peak'
    ws8.cell(row=row, column=9).value = 'Intensity (Peak-Threshold)'
    format_header_row(ws8, row, 9)
    row += 1
    
    threshold = results[station]['threshold']
    total_events = len(event_df)
    
    seasons_data = [
        ('Long Rains (Apr-Jun)', long_rains),
        ('Short Rains (Oct-Dec)', short_rains),
        ('Dry Season', dry_season)
    ]
    
    for season_name, season_df in seasons_data:
        ws8.cell(row=row, column=1).value = season_name
        ws8.cell(row=row, column=2).value = len(season_df)
        if total_events > 0:
            ws8.cell(row=row, column=3).value = f"{(len(season_df) / total_events * 100):.1f}%"
        else:
            ws8.cell(row=row, column=3).value = "0%"
        
        if len(season_df) > 0:
            ws8.cell(row=row, column=4).value = f"{season_df['duration'].mean():.2f}"
            ws8.cell(row=row, column=5).value = f"{season_df['duration'].median():.2f}"
            ws8.cell(row=row, column=6).value = f"{season_df['peak'].mean():.2f}"
            ws8.cell(row=row, column=7).value = f"{season_df['peak'].median():.2f}"
            ws8.cell(row=row, column=8).value = f"{season_df['peak'].max():.2f}"
            ws8.cell(row=row, column=9).value = f"{(season_df['peak'].mean() - threshold):.2f}"
        else:
            ws8.cell(row=row, column=4).value = "N/A"
            ws8.cell(row=row, column=5).value = "N/A"
            ws8.cell(row=row, column=6).value = "N/A"
            ws8.cell(row=row, column=7).value = "N/A"
            ws8.cell(row=row, column=8).value = "N/A"
            ws8.cell(row=row, column=9).value = "N/A"
        
        row += 1
    
    row += 2

for col in range(1, 10):
    ws8.column_dimensions[chr(64 + col)].width = 16

# ===== SHEET 9: CROSS-SEASONAL CORRELATION =====

ws9 = wb.create_sheet("9. Cross-Seasonal Correlation")
row = 1

ws9.cell(row=row, column=1).value = "ANALYSIS 9: CROSS-SEASONAL CORRELATION"
ws9.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws9.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws9.merge_cells(f'A{row}:E{row}')
row += 2

ws9.cell(row=row, column=1).value = "Analysis: How do long rains and short rains affect each other?"
ws9.cell(row=row, column=1).font = Font(italic=True, size=10)
ws9.merge_cells(f'A{row}:E{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df'].copy()
    
    ws9.cell(row=row, column=1).value = f"{station.upper()}"
    ws9.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws9.cell(row=row, column=1).fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    ws9.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Aggregate peaks by season and year
    seasonal_yearly = event_df.groupby(['year', 'season'])['peak'].max().reset_index()
    
    long_rains_yearly = seasonal_yearly[seasonal_yearly['season'] == 'Long Rains (Apr-Jun)'].set_index('year')['peak']
    short_rains_yearly = seasonal_yearly[seasonal_yearly['season'] == 'Short Rains (Oct-Dec)'].set_index('year')['peak']
    
    # Align years
    common_years = long_rains_yearly.index.intersection(short_rains_yearly.index)
    
    if len(common_years) > 1:
        lr_peaks = long_rains_yearly.loc[common_years].values
        sr_peaks = short_rains_yearly.loc[common_years].values
        
        # Correlation
        corr, pval = stats.pearsonr(lr_peaks, sr_peaks)
        
        ws9.cell(row=row, column=1).value = 'Years with Both Seasons'
        ws9.cell(row=row, column=2).value = len(common_years)
        row += 1
        
        ws9.cell(row=row, column=1).value = 'Correlation (Long Rains vs Short Rains peaks)'
        ws9.cell(row=row, column=2).value = f"{corr:.3f}"
        row += 1
        
        ws9.cell(row=row, column=1).value = 'P-value'
        ws9.cell(row=row, column=2).value = f"{pval:.4f}"
        row += 1
        
        significance = "Statistically significant" if pval < 0.05 else "Not statistically significant"
        ws9.cell(row=row, column=1).value = 'Significance (α=0.05)'
        ws9.cell(row=row, column=2).value = significance
        row += 1
        
        if corr > 0.5:
            interpretation = "Strong positive: High long rains floods often followed by high short rains floods"
        elif corr > 0.2:
            interpretation = "Moderate positive: Some relationship between seasonal flood peaks"
        elif corr > -0.2:
            interpretation = "Weak/No correlation: Seasons are relatively independent"
        elif corr > -0.5:
            interpretation = "Moderate negative: High long rains often followed by low short rains"
        else:
            interpretation = "Strong negative: High long rains inversely related to short rains"
        
        ws9.cell(row=row, column=1).value = 'Interpretation'
        ws9.cell(row=row, column=2).value = interpretation
        row += 1
    else:
        ws9.cell(row=row, column=1).value = 'Insufficient data for correlation'
        ws9.cell(row=row, column=2).value = '(Need both seasons in same years)'
        row += 1
    
    row += 2

ws9.column_dimensions['A'].width = 40
ws9.column_dimensions['B'].width = 25
ws9.column_dimensions['C'].width = 20
ws9.column_dimensions['D'].width = 20
ws9.column_dimensions['E'].width = 20

# ===== SHEET 10: SEASONAL INSIGHTS & SUMMARY =====

ws10 = wb.create_sheet("10. Seasonal Insights")
row = 1

ws10.cell(row=row, column=1).value = "ANALYSIS 10: SEASONAL INSIGHTS SUMMARY"
ws10.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws10.cell(row=row, column=1).fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ws10.merge_cells(f'A{row}:D{row}')
row += 2

for station in STATIONS:
    event_df = results[station]['event_df']
    threshold = results[station]['threshold']
    
    ws10.cell(row=row, column=1).value = f"{station.upper()} - SEASONAL SUMMARY"
    ws10.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
    ws10.cell(row=row, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws10.merge_cells(f'A{row}:D{row}')
    row += 1
    
    long_rains = event_df[event_df['season'] == 'Long Rains (Apr-Jun)']
    short_rains = event_df[event_df['season'] == 'Short Rains (Oct-Dec)']
    dry_season = event_df[event_df['season'] == 'Dry Season']
    
    # Determine which season drives flooding
    if len(long_rains) > len(short_rains) and len(long_rains) > len(dry_season):
        dominant_season = "Long Rains (Apr-Jun)"
        dominant_pct = (len(long_rains) / len(event_df) * 100)
    elif len(short_rains) > len(long_rains) and len(short_rains) > len(dry_season):
        dominant_season = "Short Rains (Oct-Dec)"
        dominant_pct = (len(short_rains) / len(event_df) * 100)
    else:
        dominant_season = "Dry Season" if len(dry_season) > 0 else "No dominant season"
        dominant_pct = (len(dry_season) / len(event_df) * 100) if len(event_df) > 0 else 0
    
    insights = []
    
    ws10.cell(row=row, column=1).value = "Key Finding:"
    ws10.cell(row=row, column=2).value = f"Dominant flood season is {dominant_season} ({dominant_pct:.1f}% of events)"
    row += 2
    
    if len(long_rains) > 0:
        lr_avg_peak = long_rains['peak'].mean()
        lr_intensity = lr_avg_peak - threshold
        ws10.cell(row=row, column=1).value = "Long Rains Intensity:"
        ws10.cell(row=row, column=2).value = f"Mean peak exceeds threshold by {lr_intensity:.0f} units"
        row += 1
    
    if len(short_rains) > 0:
        sr_avg_peak = short_rains['peak'].mean()
        sr_intensity = sr_avg_peak - threshold
        ws10.cell(row=row, column=1).value = "Short Rains Intensity:"
        ws10.cell(row=row, column=2).value = f"Mean peak exceeds threshold by {sr_intensity:.0f} units"
        row += 1
    
    if len(long_rains) > 0 and len(short_rains) > 0:
        if long_rains['peak'].mean() > short_rains['peak'].mean():
            ws10.cell(row=row, column=1).value = "Seasonal Comparison:"
            ws10.cell(row=row, column=2).value = f"Long Rains produce higher flood peaks on average"
            row += 1
        else:
            ws10.cell(row=row, column=1).value = "Seasonal Comparison:"
            ws10.cell(row=row, column=2).value = f"Short Rains produce higher flood peaks on average"
            row += 1
    
    ws10.cell(row=row, column=1).value = "Flood Frequency by Season:"
    ws10.cell(row=row, column=2).value = f"LR: {len(long_rains)} events | SR: {len(short_rains)} events | Dry: {len(dry_season)} events"
    row += 2
    
    row += 1

ws10.column_dimensions['A'].width = 35
ws10.column_dimensions['B'].width = 50
ws10.column_dimensions['C'].width = 20
ws10.column_dimensions['D'].width = 20

# ===== SAVE WORKBOOK =====

wb.save(OUTPUT_FILE)
print(f"\\n✓ Excel file saved successfully: {OUTPUT_FILE}")
print(f"\\n{'='*90}")
print(f"✓ ANALYSIS COMPLETE!")
print(f"{'='*90}")
print(f"\\nSheets created:")
print("  1. Event Frequency & Characterization")
print("  2. Decade Trends")
print("  3. Seasonality (Monthly Distribution)")
print("  4. Spatial Comparison")
print("  5. Annual Maximum & POT")
print("  6. Long Rains Analysis (Apr-Jun)")
print("  7. Short Rains Analysis (Oct-Dec)")
print("  8. Seasonal Comparison - Which Season Drives Flooding?")
print("  9. Cross-Seasonal Correlation - How Seasons Relate to Each Other")
print("  10. Seasonal Insights Summary")
print(f"\\n{'='*90}")
print(f"Output file: {OUTPUT_FILE}")
print(f"{'='*90}")
