"""
RQ1: Flow Duration Curves, GEV Return Periods,
Rolling Statistics, and Extreme Value Analysis

Data NOTE: The dataset only contains Long Rains (Mar-May) records.
(OND-AMJ correlation, dry season stats) cannot be
done without Oct-Dec data.

"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from scipy.stats import genextreme
import warnings
import os

warnings.filterwarnings('ignore')

# Excel Config

EXCEL_FILE = 'TanaRiver Flow.xlsx'
SHEET_NAME = 'Rawdata'
OUTPUT_DIR = './rq1a_analysis/'
os.makedirs(OUTPUT_DIR, exist_ok=True)

THRESHOLDS = {'Bura': 721, 'Galole': 723, 'Garsen': 1723}
STATIONS = ['Bura', 'Galole', 'Garsen']
STATION_COLORS = {'Bura': '#1f77b4', 'Galole': '#ff7f0e', 'Garsen': '#2ca02c'}

sns.set_style("whitegrid")
plt.rcParams.update({
    'figure.figsize': (14, 8),
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.labelsize': 12,
})

raw = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME, header=1, usecols=[0, 1, 2, 3])
raw.columns = ['date', 'Bura', 'Galole', 'Garsen']
raw['date'] = pd.to_datetime(raw['date'], errors='coerce')
for col in STATIONS:
    raw[col] = pd.to_numeric(raw[col], errors='coerce')
raw['year'] = raw['date'].dt.year
raw['month'] = raw['date'].dt.month
raw['doy'] = raw['date'].dt.dayofyear
raw = raw.dropna(subset=['date']).reset_index(drop=True)

print(f"Data loaded: {raw['date'].min().date()} to {raw['date'].max().date()}")
print(f"Records: {len(raw)}, Years: {raw['year'].min()}-{raw['year'].max()}")
print(f"Months present: {sorted(raw['month'].unique())} (Long Rains season only)")
print()

# MEAN ANNUAL CYCLE (Long Rains)
print("SECTION 1: SEASONAL CHARACTERISTICS (Long Rains: Mar-May)")

fig, axes = plt.subplots(3, 1, figsize=(14, 12), sharex=True)
fig.suptitle('Mean Daily River Level During Long Rains Season (Mar-May)\n'
             'With 14-day Rolling Mean and Flood Threshold',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    ax = axes[idx]
    daily_by_doy = raw.groupby('doy')[station].agg(['mean', 'std']).reset_index()

    doys = daily_by_doy['doy'].values
    means = daily_by_doy['mean'].values
    stds = daily_by_doy['std'].values

    ax.fill_between(doys, means - stds, means + stds, alpha=0.2, color=STATION_COLORS[station])
    ax.plot(doys, means, '-', color=STATION_COLORS[station], linewidth=2.5,
            label='Climatological mean +/- std')

    # 14day rolling mean
    rolling_mean = pd.Series(means).rolling(14, min_periods=1, center=True).mean()
    ax.plot(doys, rolling_mean.values, '--', color='navy', linewidth=2,
            label='14-day smoothed')

    ax.axhline(THRESHOLDS[station], color='red', linestyle='--', linewidth=2,
               alpha=0.8, label=f'Flood threshold: {THRESHOLDS[station]}')

    ax.set_ylabel('River Level', fontweight='bold')
    ax.set_title(f'{station} Station', fontweight='bold')
    ax.legend(loc='upper right', fontsize=9)
    ax.grid(alpha=0.3)

axes[-1].set_xlabel('Day of Year (60=Mar 1, 120=Apr 30, 151=May 31)', fontweight='bold')
axes[-1].set_xticks([60, 70, 80, 90, 100, 110, 120, 130, 140, 151])
axes[-1].set_xticklabels(['Mar 1', '', 'Mar 21', '', 'Apr 10', '', 'Apr 30', '', 'May 20', 'May 31'])
plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}01_Mean_Annual_Cycle.png', dpi=300, bbox_inches='tight')
print("Saved: 01_Mean_Annual_Cycle.png")
plt.close()

# Monthly stats
for station in STATIONS:
    for month in [3, 4, 5]:
        mname = {3: 'March', 4: 'April', 5: 'May'}[month]
        s = raw[raw['month'] == month][station].dropna()
        pct_above = (s > THRESHOLDS[station]).sum() / len(s) * 100 if len(s) > 0 else 0
        print(f"  {station} {mname:5s}: mean={s.mean():.1f}, std={s.std():.1f}, "
              f"max={s.max():.1f}, days above threshold={pct_above:.1f}%")

# FLOW DURATION CURVES (FDCs)

fig, axes = plt.subplots(1, 3, figsize=(16, 6))
fig.suptitle('Flow Duration Curves - Long Rains Season (Mar-May)\n'
             'Percentage of Time River Level Exceeds a Given Value',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    ax = axes[idx]
    data = raw[station].dropna().values
    sorted_data = np.sort(data)[::-1]
    n = len(sorted_data)
    exceedance_prob = np.arange(1, n + 1) / n * 100

    ax.plot(sorted_data, exceedance_prob, color=STATION_COLORS[station], linewidth=2.5)

    # Threshold
    pct_at_threshold = (data > THRESHOLDS[station]).sum() / n * 100
    ax.scatter([THRESHOLDS[station]], [pct_at_threshold], color='red', s=120, zorder=5,
               edgecolors='black', linewidth=1.5,
               label=f'Threshold ({THRESHOLDS[station]}): {pct_at_threshold:.1f}%')

    # Reference lines
    for pct, lvl in [(5, np.percentile(data, 95)), (10, np.percentile(data, 90)),
                     (25, np.percentile(data, 75))]:
        ax.axhline(pct, color='gray', linestyle=':', linewidth=0.5, alpha=0.5)
        ax.scatter([lvl], [pct], color='gray', s=30, zorder=5)
        ax.annotate(f'{pct}%: {lvl:.0f}', xy=(lvl, pct), xytext=(10, 0),
                    textcoords='offset points', fontsize=8, color='gray', va='center')

    ax.set_xlabel('River Level', fontweight='bold')
    ax.set_ylabel('Exceedance Probability (%)', fontweight='bold')
    ax.set_title(f'{station} Station (n={n})', fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(alpha=0.3)
    ax.set_xlim(left=0)

    print(f"  {station}: Threshold {THRESHOLDS[station]} exceeded {pct_at_threshold:.2f}% of the Long Rains season")
    print(f"    5% exceedance level: {np.percentile(data, 95):.1f}")
    print(f"    10% exceedance level: {np.percentile(data, 90):.1f}")
    print(f"    25% exceedance level: {np.percentile(data, 75):.1f}")
    print(f"    Median (50%): {np.percentile(data, 50):.1f}")

plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}02_Flow_Duration_Curves.png', dpi=300, bbox_inches='tight')
print("Saved: 02_Flow_Duration_Curves.png")
plt.close()
# 3. GEV RETURN PERIOD ANALYSIS

print("  (Using annual maxima within Long Rains season)")
def gev_return_level(xi, sigma, mu, T):
    p = 1.0 - 1.0 / T
    if abs(xi) < 1e-12:
        return mu - sigma * np.log(-np.log(p))
    return mu + (sigma / xi) * ((-np.log(p)) ** (-xi) - 1.0)

def fit_gev(data):
    data = data[~np.isnan(data)]
    c, loc, scale = stats.genextreme.fit(data)
    xi = -c
    # Sanity check: if shape parameter is extreme (>2), the fit is unreliable
    reliable = abs(xi) < 2.0
    return {'xi': xi, 'sigma': scale, 'mu': loc, 'n': len(data), 'reliable': reliable}

return_periods = np.array([1.5, 2, 3, 5, 10, 20, 50, 100])
gev_results = {}

fig, axes = plt.subplots(2, 3, figsize=(18, 12))
fig.suptitle('GEV Return Level Analysis - Annual Maxima (Long Rains Season)',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    # Annual maxima within Long Rains season
    annual_max = raw.groupby('year')[station].max().dropna()
    am_vals = annual_max.values

    # Fit GEV
    params = fit_gev(am_vals)
    gev_results[station] = params

    print(f"\n  {station} Station (N={params['n']} years):")
    print(f"    GEV shape (xi): {params['xi']:.4f}  "
          f"({'heavy tail' if params['xi'] > 0 else 'light tail' if params['xi'] < 0 else 'Fréchet'})")
    print(f"    GEV scale (sigma): {params['sigma']:.4f}")
    print(f"    GEV location (mu): {params['mu']:.4f}")
    if not params['reliable']:
        print(f"    *** WARNING: Shape parameter xi={params['xi']:.2f} is extreme (>2).")
        print(f"        GEV fit is UNRELIABLE for this station. Return levels beyond")
        print(f"        2-year should be treated with extreme caution.")

    # Return levels
    rls = [gev_return_level(params['xi'], params['sigma'], params['mu'], T)
           for T in return_periods]

    #  Plot 1: Return level plot 
    ax = axes[0, idx]
    sorted_am = np.sort(am_vals)[::-1]
    n = len(sorted_am)
    empirical_rp = (n + 1) / np.arange(1, n + 1)

    ax.scatter(empirical_rp, sorted_am, c=STATION_COLORS[station], s=70,
               alpha=0.7, edgecolors='black', zorder=5, label='Observed annual max')
    ax.plot(return_periods, rls, 'r-o', linewidth=2.5, markersize=7, label='GEV fit')

    # Confidence band via bootstrap (only for reliable fits)
    if params['reliable']:
        rng = np.random.default_rng(42)
        n_boot = 100
        boot_rls = {T: [] for T in return_periods}
        for _ in range(n_boot):
            sample = rng.choice(am_vals, size=len(am_vals), replace=True)
            try:
                bp = fit_gev(sample)
                if bp['reliable']:
                    for T in return_periods:
                        boot_rls[T].append(gev_return_level(bp['xi'], bp['sigma'], bp['mu'], T))
            except Exception:
                pass
        if boot_rls[return_periods[0]]:
            lower = [np.percentile(boot_rls[T], 5) for T in return_periods]
            upper = [np.percentile(boot_rls[T], 95) for T in return_periods]
            ax.fill_between(return_periods, lower, upper, alpha=0.2, color='red', label='90% CI (bootstrap)')
    else:
        ax.text(0.05, 0.05, 'GEV fit unreliable - no bootstrap CI', transform=ax.transAxes,
                fontsize=10, color='red', style='italic')

    ax.axhline(THRESHOLDS[station], color='orange', linestyle='--', linewidth=2,
               alpha=0.7, label=f'Threshold: {THRESHOLDS[station]}')

    ax.set_xscale('log')
    ax.set_xlabel('Return Period (years, log scale)', fontweight='bold')
    ax.set_ylabel('River Level (annual max)', fontweight='bold')
    title = f'{station} - Return Level Plot'
    if not params['reliable']:
        title += '\n[UNRELIABLE GEV FIT]'
    ax.set_title(title, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3, which='both')

    # Print return levels
    print(f"    Return levels:")
    for T, rl in zip(return_periods, rls):
        if params['reliable'] or T <= 3:
            print(f"      {T:5.1f}-year: {rl:.1f}")
        else:
            print(f"      {T:5.1f}-year: {rl:.1f}  (unreliable)")

    # Find return period for threshold
    for T_check in np.linspace(1.0, 20.0, 500):
        rl_check = gev_return_level(params['xi'], params['sigma'], params['mu'], T_check)
        if rl_check >= THRESHOLDS[station]:
            print(f"    Threshold {THRESHOLDS[station]} ~ {T_check:.2f}-year return period")
            break

    #  Plot 2: GEV PDF fit 
    ax2 = axes[1, idx]
    x_range = np.linspace(max(0, params['mu'] - 3 * params['sigma']),
                          params['mu'] + 5 * params['sigma'], 200)

    c_scipy = -params['xi']
    gev_pdf = genextreme.pdf(x_range, c_scipy, loc=params['mu'], scale=params['sigma'])
    ax2.hist(am_vals, bins=min(15, len(am_vals)), density=True,
             color=STATION_COLORS[station], alpha=0.5, edgecolor='black',
             label='Observed annual max')
    ax2.plot(x_range, gev_pdf, 'r-', linewidth=2.5, label='GEV fit')

    # Mark threshold and return levels
    ax2.axvline(THRESHOLDS[station], color='orange', linestyle='--', linewidth=2,
                label=f'Threshold: {THRESHOLDS[station]}')
    ax2.axvline(params['mu'], color='green', linestyle=':', linewidth=1.5,
                label=f'Mu (location): {params["mu"]:.0f}')

    ax2.set_xlabel('River Level (annual max)', fontweight='bold')
    ax2.set_ylabel('Density', fontweight='bold')
    ax2.set_title(f'{station} - GEV PDF Fit', fontweight='bold')
    ax2.legend(fontsize=9)
    ax2.grid(alpha=0.3)

plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}03_GEV_Return_Periods.png', dpi=300, bbox_inches='tight')
print("\n Saved: 03_GEV_Return_Periods.png")
plt.close()

# ENHANCED TIME SERIES rolling stats

fig, axes = plt.subplots(3, 1, figsize=(16, 12), sharex=True)
fig.suptitle('River Level Time Series with Rolling Statistics and Flood Threshold\n'
             '(Long Rains Season: Mar-May)',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    ax = axes[idx]

    ax.plot(raw['date'], raw[station], color=STATION_COLORS[station],
            linewidth=0.6, alpha=0.5, label='Daily level')

    # 7-day rolling mean
    rolling_7 = raw[station].rolling(7, min_periods=1).mean()
    ax.plot(raw['date'], rolling_7, color='steelblue', linewidth=1.2, alpha=0.8,
            label='7-day rolling mean')

    # 14-day rolling max
    rolling_14 = raw[station].rolling(14, min_periods=1).max()
    ax.plot(raw['date'], rolling_14, color='darkorange', linewidth=1.5, alpha=0.8,
            label='14-day rolling max')

    # 30-day rolling mean
    rolling_30 = raw[station].rolling(30, min_periods=1).mean()
    ax.plot(raw['date'], rolling_30, color='darkgreen', linewidth=1.8, alpha=0.8,
            label='30-day rolling mean')

    # Threshold
    ax.axhline(THRESHOLDS[station], color='red', linestyle='--', linewidth=2,
               alpha=0.8, label=f'Flood threshold: {THRESHOLDS[station]}')

    # Mark years that exceeded threshold
    above = raw[raw[station] > THRESHOLDS[station]]
    ax.scatter(above['date'], above[station], color='red', s=6, alpha=0.3, zorder=5)

    # Add year labels
    for yr in range(raw['year'].min(), raw['year'].max() + 1, 5):
        yr_start = raw[raw['year'] == yr]['date'].min()
        ax.axvline(yr_start, color='gray', linestyle=':', linewidth=0.5, alpha=0.3)

    ax.set_ylabel('River Level', fontweight='bold')
    ax.set_title(f'{station} Station (Threshold: {THRESHOLDS[station]})', fontweight='bold')
    ax.legend(loc='upper left', fontsize=9, ncol=2)
    ax.grid(alpha=0.3)

axes[-1].set_xlabel('Date', fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}04_Time_Series_Enhanced.png', dpi=300, bbox_inches='tight')
print("[OK] Saved: 04_Time_Series_Enhanced.png")
plt.close()

# INTER-STATION CROSS-CORRELATION & LAG

fig, axes = plt.subplots(1, 3, figsize=(16, 5))
fig.suptitle('Cross-correlation Between Stations (Lag Analysis)',
             fontsize=14, fontweight='bold')

station_pairs = [('Bura', 'Galole'), ('Bura', 'Garsen'), ('Galole', 'Garsen')]

for idx, (s_up, s_down) in enumerate(station_pairs):
    ax = axes[idx]
    up = raw[s_up].dropna().values
    down = raw[s_down].dropna().values
    n = min(len(up), len(down))
    up = up[:n]
    down = down[:n]

    max_lag = 15
    lags = range(-max_lag, max_lag + 1)
    correlations = []

    for lag in lags:
        if lag >= 0:
            a, b = up[:n - lag] if lag > 0 else up, down[lag:]
        else:
            a, b = up[-lag:], down[:n + lag]
        if len(a) > 1 and len(b) > 1:
            c = np.corrcoef(a[:min(len(a), len(b))], b[:min(len(a), len(b))])[0, 1]
        else:
            c = 0
        correlations.append(c)

    ax.plot(list(lags), correlations, 'o-', color='steelblue', markersize=4)
    ax.axvline(0, color='red', linestyle='--', linewidth=1, alpha=0.5)
    ax.axhline(0, color='gray', linestyle=':', linewidth=0.5)

    max_corr_idx = np.argmax(correlations)
    best_lag = list(lags)[max_corr_idx]
    best_corr = correlations[max_corr_idx]

    ax.axvline(best_lag, color='green', linestyle='--', linewidth=2,
               label=f'Best lag: {best_lag}d (r={best_corr:.3f})')

    ax.set_xlabel('Lag (days, positive = downstream delayed)', fontweight='bold')
    ax.set_ylabel('Cross-correlation', fontweight='bold')
    ax.set_title(f'{s_up} -> {s_down}', fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)

    print(f"  {s_up} -> {s_down}: Best lag = {best_lag} days, r = {best_corr:.3f}")

plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}05_Interstation_Lag.png', dpi=300, bbox_inches='tight')
print("[OK] Saved: 05_Interstation_Lag.png")
plt.close()

# YEAR-YEAR ANNUAL MAXIMA & THRESHOLD EXCEEDANCE

fig, axes = plt.subplots(3, 1, figsize=(16, 12), sharex=True)
fig.suptitle('Annual Maximum River Level and Days Above Threshold\n(Long Rains Season)',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    ax1 = axes[idx]
    ax2 = ax1.twinx()

    annual_max = raw.groupby('year')[station].max()
    days_above = raw.groupby('year')[station].apply(lambda x: (x > THRESHOLDS[station]).sum())

    # Bar: days above threshold
    bars = ax2.bar(annual_max.index, days_above.values, alpha=0.3, color='coral',
                   label='Days above threshold', width=0.8)

    # Line: annual max
    ax1.plot(annual_max.index, annual_max.values, 'o-', color=STATION_COLORS[station],
             linewidth=2, markersize=5, label='Annual max')

    # Threshold
    ax1.axhline(THRESHOLDS[station], color='red', linestyle='--', linewidth=2, alpha=0.7)

    # Trend line
    valid = annual_max.dropna()
    if len(valid) > 2:
        z = np.polyfit(valid.index, valid.values, 1)
        p = np.poly1d(z)
        ax1.plot(valid.index, p(valid.index), 'k--', linewidth=1.5, alpha=0.5,
                 label=f'Trend: {z[0]:+.2f}/yr')

    ax1.set_ylabel('Annual Max River Level', fontweight='bold', color=STATION_COLORS[station])
    ax2.set_ylabel('Days Above Threshold', fontweight='bold', color='coral')
    ax1.set_title(f'{station} Station', fontweight='bold')
    ax1.legend(loc='upper left', fontsize=9)
    ax2.legend(loc='upper right', fontsize=9)
    ax1.grid(alpha=0.3)

    # Stats
    print(f"  {station}: Mean annual max={annual_max.mean():.1f}, "
          f"Std={annual_max.std():.1f}, "
          f"Mean days above threshold={days_above.mean():.1f}")

axes[-1].set_xlabel('Year', fontweight='bold')
plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}06_Annual_Maxima_Events.png', dpi=300, bbox_inches='tight')
print("[OK] Saved: 06_Annual_Maxima_Events.png")
plt.close()

# MONTHLY DISTRIBUTION WITHIN LONG RAINS

fig, axes = plt.subplots(1, 3, figsize=(16, 6))
fig.suptitle('Distribution of River Levels by Month (Long Rains: Mar-May)',
             fontsize=14, fontweight='bold')

for idx, station in enumerate(STATIONS):
    ax = axes[idx]
    month_data = [raw[raw['month'] == m][station].dropna().values for m in [3, 4, 5]]

    bp = ax.boxplot(month_data, labels=['March', 'April', 'May'],
                     patch_artist=True, widths=0.6,
                     medianprops=dict(color='black', linewidth=2))

    colors_m = ['#3498db', '#2ecc71', '#e74c3c']
    for patch, color in zip(bp['boxes'], colors_m):
        patch.set_facecolor(color)
        patch.set_alpha(0.6)

    ax.axhline(THRESHOLDS[station], color='red', linestyle='--', linewidth=2,
               label=f'Threshold: {THRESHOLDS[station]}')

    # Count above threshold per month
    for m_idx, m in enumerate([3, 4, 5]):
        above = (raw[raw['month'] == m][station] > THRESHOLDS[station]).sum()
        total = len(raw[raw['month'] == m][station].dropna())
        pct = above / total * 100 if total > 0 else 0
        ax.text(m_idx + 1, ax.get_ylim()[1] * 0.95,
                f'{above}/{total}\n({pct:.0f}%)',
                ha='center', va='top', fontsize=9, fontweight='bold',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

    ax.set_ylabel('River Level', fontweight='bold')
    ax.set_title(f'{station} Station', fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(axis='y', alpha=0.3)

plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}07_Monthly_Distribution.png', dpi=300, bbox_inches='tight')
print("[OK] Saved: 07_Monthly_Distribution.png")
plt.close()

# JOINT SCATTER: BURA vs GALOLE vs GARSEN

fig, axes = plt.subplots(1, 3, figsize=(16, 5))
fig.suptitle('Inter-station Scatter: How River Levels Co-vary',
             fontsize=14, fontweight='bold')

for idx, (s1, s2) in enumerate([('Bura', 'Galole'), ('Bura', 'Garsen'), ('Galole', 'Garsen')]):
    ax = axes[idx]
    x = raw[s1].dropna()
    y = raw[s2].dropna()
    common_idx = x.index.intersection(y.index)
    x_vals = x.loc[common_idx].values
    y_vals = y.loc[common_idx].values

    ax.scatter(x_vals, y_vals, alpha=0.3, s=15, c='steelblue')

    # Regression
    slope, intercept, r, p, se = stats.linregress(x_vals, y_vals)
    x_line = np.linspace(x_vals.min(), x_vals.max(), 100)
    ax.plot(x_line, slope * x_line + intercept, 'r-', linewidth=2,
            label=f'r={r:.3f}, y={slope:.2f}x+{intercept:.0f}')

    ax.set_xlabel(s1, fontweight='bold')
    ax.set_ylabel(s2, fontweight='bold')
    ax.set_title(f'{s1} vs {s2}', fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(alpha=0.3)

    print(f"  {s1} vs {s2}: r={r:.3f}, slope={slope:.3f}, intercept={intercept:.1f}")

plt.tight_layout()
plt.savefig(f'{OUTPUT_DIR}08_Interstation_Scatter.png', dpi=300, bbox_inches='tight')
print("[OK] Saved: 08_Interstation_Scatter.png")
plt.close()


# COMPREHENSIVE SUMMARY STATISTICS

summary_data = []
for station in STATIONS:
    am = raw.groupby('year')[station].max().dropna()
    params = gev_results[station]
    rls = {T: gev_return_level(params['xi'], params['sigma'], params['mu'], T)
           for T in [2, 5, 10, 20]}

    # Find threshold return period
    threshold_rp = None
    for T_check in np.linspace(1.0, 20.0, 500):
        rl_check = gev_return_level(params['xi'], params['sigma'], params['mu'], T_check)
        if rl_check >= THRESHOLDS[station]:
            threshold_rp = T_check
            break

    summary_data.append({
        'Station': station,
        'N_years': params['n'],
        'Mean Annual Max': f"{am.mean():.1f}",
        'Std Annual Max': f"{am.std():.1f}",
        'Min Annual Max': f"{am.min():.1f}",
        'Max Annual Max': f"{am.max():.1f}",
        'GEV xi': f"{params['xi']:.4f}",
        'GEV sigma': f"{params['sigma']:.2f}",
        'GEV mu': f"{params['mu']:.2f}",
        'Threshold': THRESHOLDS[station],
        'Threshold RP (yr)': f"{threshold_rp:.2f}" if threshold_rp else "N/A",
        '2-yr RL': f"{rls[2]:.1f}",
        '5-yr RL': f"{rls[5]:.1f}",
        '10-yr RL': f"{rls[10]:.1f}",
        '20-yr RL': f"{rls[20]:.1f}",
    })

summary_df = pd.DataFrame(summary_data)
print(summary_df.to_string(index=False))

print("EXPORTING TO EXCEL")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

def write_sheet(ws, title, headers, rows):
    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=1, column=1).fill = header_fill
    end_col = chr(64 + min(len(headers), 26))
    ws.merge_cells(f'A1:{end_col}1')
    ws.append([])
    for col_idx, h in enumerate(headers[:26], 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True)
    for row_idx, row in enumerate(rows, 4):
        for col_idx, val in enumerate(row[:26], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

# Sheet 1: Summary
ws1 = wb.create_sheet("Summary")
sum_headers = list(summary_df.columns)
write_sheet(ws1, "Extreme Value Summary", sum_headers,
            [list(r) for _, r in summary_df.iterrows()])

# Sheet 2: GEV Return Levels
ws2 = wb.create_sheet("GEV Return Levels")
gev_headers = ['Station', 'Return Period', 'Return Level']
gev_rows = []
for station in STATIONS:
    params = gev_results[station]
    for T in [1.5, 2, 3, 5, 10, 20, 50, 100]:
        rl = gev_return_level(params['xi'], params['sigma'], params['mu'], T)
        gev_rows.append([station, T, f"{rl:.2f}"])
write_sheet(ws2, "GEV Return Levels by Station", gev_headers, gev_rows)

# Sheet 3: Annual Maxima Time Series
ws3 = wb.create_sheet("Annual Maxima")
am_headers = ['Year'] + [f'{s} Annual Max' for s in STATIONS] + [f'{s} Days > Threshold' for s in STATIONS]
am_rows = []
for yr in sorted(raw['year'].unique()):
    row = [yr]
    for s in STATIONS:
        am_val = raw[raw['year'] == yr][s].max()
        row.append(f"{am_val:.2f}" if not pd.isna(am_val) else "")
    for s in STATIONS:
        days = (raw[raw['year'] == yr][s] > THRESHOLDS[s]).sum()
        row.append(days)
    am_rows.append(row)
write_sheet(ws3, "Annual Maxima Time Series", am_headers, am_rows)

# Sheet 4: FDC Exceedance Table
ws4 = wb.create_sheet("FDC Exceedance")
fdc_headers = ['Station', 'Threshold', 'Days in Season', 'Days > Threshold',
               'Exceedance %', '5% Level', '10% Level', '25% Level', 'Median']
fdc_rows = []
for station in STATIONS:
    data = raw[station].dropna()
    total = len(data)
    above = (data > THRESHOLDS[station]).sum()
    pct = above / total * 100
    fdc_rows.append([
        station, THRESHOLDS[station], total, above, f"{pct:.2f}",
        f"{np.percentile(data, 95):.1f}", f"{np.percentile(data, 90):.1f}",
        f"{np.percentile(data, 75):.1f}", f"{np.percentile(data, 50):.1f}"
    ])
write_sheet(ws4, "Flow Duration Curve Exceedance", fdc_headers, fdc_rows)

# Sheet 5: Inter-station Lags
ws5 = wb.create_sheet("Inter-station Lags")
lag_headers = ['From', 'To', 'Best Lag (days)', 'Correlation']
lag_rows = []
for s_up, s_down in station_pairs:
    up = raw[s_up].dropna().values
    down = raw[s_down].dropna().values
    n = min(len(up), len(down))
    best_lag, best_corr = 0, -1
    for lag in range(-15, 16):
        if lag >= 0:
            a, b = up[:n - lag] if lag > 0 else up, down[lag:]
        else:
            a, b = up[-lag:], down[:n + lag]
        if len(a) > 1 and len(b) > 1:
            c = np.corrcoef(a[:min(len(a), len(b))], b[:min(len(a), len(b))])[0, 1]
        else:
            c = 0
        if c > best_corr:
            best_corr = c
            best_lag = lag
    lag_rows.append([s_up, s_down, best_lag, f"{best_corr:.4f}"])
write_sheet(ws5, "Inter-station Lag Analysis", lag_headers, lag_rows)

for ws in wb.worksheets:
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 18

output_file = 'RQ1a_Rainfall_GEV_Analysis.xlsx'
wb.save(output_file)
print(f"[OK] Excel saved: {output_file}")
# SUMMARY
print("\n" + "=" * 80)
print("RQ1a ANALYSIS COMPLETE!")
print("=" * 80)
print(f"\nOutputs:")
print(f"  Visualizations: {OUTPUT_DIR}")
print(f"  Excel: {output_file}")
print(f"\nKey findings:")
for station in STATIONS:
    params = gev_results[station]
    rl2 = gev_return_level(params['xi'], params['sigma'], params['mu'], 2)
    rl5 = gev_return_level(params['xi'], params['sigma'], params['mu'], 5)
    data = raw[station].dropna()
    pct = (data > THRESHOLDS[station]).sum() / len(data) * 100
    print(f"  {station}:")
    print(f"    - Threshold {THRESHOLDS[station]} exceeded {pct:.1f}% of Long Rains season days")
    print(f"    - GEV: xi={params['xi']:.3f}, sigma={params['sigma']:.1f}, mu={params['mu']:.1f}")
    print(f"    - 2-year RL={rl2:.0f}, 5-year RL={rl5:.0f}")
print()
